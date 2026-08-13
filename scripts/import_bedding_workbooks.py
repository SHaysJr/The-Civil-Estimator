"""Import the normalized Mississippi and Tennessee pipe bedding workbooks.

The workbooks are source data only. This script snapshots their calculated rows
into SQLite so the estimator does not need Excel at runtime.
"""
from pathlib import Path
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from db import connect, init_db


def rows_from_workbook(path: Path, state: str):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb['Bedding Data']
    headers = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
    idx = {name: i for i, name in enumerate(headers) if name}

    for values in ws.iter_rows(min_row=5, values_only=True):
        jurisdiction = values[idx['Jurisdiction']]
        if not jurisdiction:
            continue
        confidence = values[idx['Data Confidence']] if 'Data Confidence' in idx else 'Confirmed / source-listed'
        source_key = 'Source Document / Section'
        bedding_class = values[idx['Bedding Class/Type']] or ''
        prohibited = 'PROHIBITED' in bedding_class.upper()
        yield {
            'state': state,
            'jurisdiction': jurisdiction,
            'system_type': values[idx['System Type']],
            'pipe_material': values[idx['Pipe Material']],
            'nominal_diameter_in': float(values[idx['Nominal Diameter (in)']]) if isinstance(values[idx['Nominal Diameter (in)']], (int, float)) else 0.0,
            'applies_all_sizes': 0 if isinstance(values[idx['Nominal Diameter (in)']], (int, float)) else 1,
            'pipe_od_in': float(values[idx['Pipe OD (in)']] or 0) if isinstance(values[idx['Pipe OD (in)']], (int, float)) else 0.0,
            'bedding_class': bedding_class,
            'depth_below_in': float(values[idx['Bedding Depth Below Pipe (in)']] or 0),
            'haunch_to_springline': 1 if str(values[idx["Haunch to Springline Req'd (Y/N)"]] or '').upper() == 'Y' else 0,
            'envelope_above_crown_in': float(values[idx['Envelope Height Above Crown (in)']] or 0),
            'trench_side_clearance_in': float(values[idx['Trench Side Clearance Each Side (in)']] or 0),
            'trench_width_in': float(values[idx['Trench Width (in)']] or 0),
            'bedding_material_spec': values[idx['Bedding Material Spec']] or '',
            'compaction_requirement': values[idx['Compaction Requirement']] or '',
            'bedding_volume_per_lf_cf': 0.0 if prohibited else float(values[idx['Bedding Volume per LF (CF)']] or 0),
            'bedding_volume_per_100_lf_cy': 0.0 if prohibited else float(values[idx['Bedding Volume per 100 LF (CY)']] or 0),
            'data_confidence': confidence or '',
            'source_document': values[idx[source_key]] or '',
            'prohibited': 1 if prohibited else 0,
        }


def import_files(files):
    init_db()
    sql = '''INSERT INTO bedding_rules(
        state,jurisdiction,system_type,pipe_material,nominal_diameter_in,applies_all_sizes,pipe_od_in,bedding_class,
        depth_below_in,haunch_to_springline,envelope_above_crown_in,trench_side_clearance_in,trench_width_in,
        bedding_material_spec,compaction_requirement,bedding_volume_per_lf_cf,bedding_volume_per_100_lf_cy,
        data_confidence,source_document,prohibited)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(state,jurisdiction,system_type,pipe_material,nominal_diameter_in,bedding_class) DO UPDATE SET
        pipe_od_in=excluded.pipe_od_in,depth_below_in=excluded.depth_below_in,
        haunch_to_springline=excluded.haunch_to_springline,envelope_above_crown_in=excluded.envelope_above_crown_in,
        trench_side_clearance_in=excluded.trench_side_clearance_in,trench_width_in=excluded.trench_width_in,
        bedding_material_spec=excluded.bedding_material_spec,compaction_requirement=excluded.compaction_requirement,
        bedding_volume_per_lf_cf=excluded.bedding_volume_per_lf_cf,
        bedding_volume_per_100_lf_cy=excluded.bedding_volume_per_100_lf_cy,
        data_confidence=excluded.data_confidence,source_document=excluded.source_document,prohibited=excluded.prohibited'''
    count = 0
    with connect() as conn:
        for path, state in files:
            for r in rows_from_workbook(path, state):
                conn.execute(sql, tuple(r.values()))
                count += 1
        conn.commit()
    return count


if __name__ == '__main__':
    args = sys.argv[1:]
    if args:
        files=[]
        for p in args:
            path=Path(p)
            state='MS' if path.name.upper().startswith('MS_') else 'TN'
            files.append((path,state))
    else:
        files=[
            (ROOT.parent / 'MS_Pipe_Bedding_Estimating_Data.xlsx','MS'),
            (ROOT.parent / 'TN_Pipe_Bedding_Estimating_Data.xlsx','TN'),
        ]
    existing=[x for x in files if x[0].exists()]
    if not existing:
        raise SystemExit('No bedding workbooks found. Pass workbook paths on the command line.')
    print(f'Imported {import_files(existing)} bedding-rule rows.')
